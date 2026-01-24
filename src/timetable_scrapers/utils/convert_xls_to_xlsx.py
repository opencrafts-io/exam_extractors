#!/usr/bin/env python3
from datetime import datetime
from pathlib import Path
import sys
from typing import Literal

import xlrd
from openpyxl import Workbook


def _get_cell_value(cell: xlrd.sheet.Cell, datemode: int):
    cell_type = cell.ctype
    value = cell.value

    if cell_type == xlrd.XL_CELL_DATE:
        mode: Literal[0, 1] = 1 if datemode == 1 else 0
        date_tuple = xlrd.xldate_as_tuple(float(value), mode)
        return datetime(*date_tuple)
    if cell_type == xlrd.XL_CELL_EMPTY:
        return None
    if cell_type == xlrd.XL_CELL_BOOLEAN:
        return bool(value)
    return value


def convert_xls_to_xlsx(xls_path: str | Path, xlsx_path: str | Path | None = None) -> Path:
    source = Path(xls_path)
    target = Path(xlsx_path) if xlsx_path else source.with_suffix(".xlsx")

    rb = xlrd.open_workbook(str(source), formatting_info=False)
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for sheet_index in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_index)
        ws = wb.create_sheet(title=rs.name)

        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                cell = rs.cell(row_idx, col_idx)
                val = _get_cell_value(cell, rb.datemode)
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=val)

    wb.save(target)
    return target


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(1)

    xls_file = sys.argv[1]
    xlsx_file = sys.argv[2] if len(sys.argv) > 2 else None

    convert_xls_to_xlsx(xls_file, xlsx_file)
