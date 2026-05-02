from typing import Any, Dict, List
from openpyxl import load_workbook

from ...base.scraper import BaseTimetableScraper
from ...registry import ScraperRegistry
from ...schemas import CourseEntry
from ...utils.time_parser import parse_exam_datetime


@ScraperRegistry.register("nursing_exams")
class NursingExamScraper(BaseTimetableScraper):
    """
    Scraper for Daystar University Nursing timetable Format.
    Matches the minimal Professor API contract.
    """

    @property
    def institution_name(self) -> str:
        return "Daystar University Nursing School"

    @property
    def timezone(self) -> str:
        return "EAT"  # East Africa Time

    def extract(self, file) -> List[CourseEntry]:
        """
        Extract timetable data from nursing excel file
        """
        column_headers = [
            "Day",
            "Campus",
            "Coordinator",
            "Courses",
            "Hours",
            "Venue",
            "Invigilators",
            "Courses_Afternoon",
            "Hours_Afternoon",
            "Venue_Afternoon",
            "Invigilators_Afternoon",
        ]

        wb_obj = load_workbook(file)
        sheet = wb_obj.active

        column_data_dict = self._read_columns(sheet, column_headers)
        morning_exams = self._extract_course_info(
            column_data_dict,
            "Courses",
            "8:30AM", "11:30AM"
        )
        afternoon_exams = self._extract_course_info(
            column_data_dict,
            "Courses_Afternoon",
            "1:30PM", "4:30PM"
        )

        courses = morning_exams + afternoon_exams
        return self.normalize_output(courses)

    def _read_columns(
        self, sheet, column_headers: List[str]
    ) -> Dict[str, List[Any]]:
        """
        Read Excel columns and store data in dictionary.
        Handles merged cells by carrying forward the last non-None value.
        """
        column_data_dict = {}
        no_carry_headers = {"Courses", "Courses_Afternoon"}

        for i, column in enumerate(
            sheet.iter_cols(values_only=True) if sheet else []
        ):
            if i >= len(column_headers):
                continue

            header = column_headers[i]
            last_value = None
            column_data = []

            for cell in column:
                if cell is not None:
                    last_value = cell
                    column_data.append(cell)
                else:
                    if header in no_carry_headers:
                        column_data.append(None)
                    else:
                        column_data.append(last_value)

            column_data_dict[header] = column_data

        return column_data_dict

    def _extract_course_info(
        self,
        column_data_dict: Dict[str, List[Any]],
        time_key: str,
        start_time_str: str,
        end_time_str: str,
    ) -> List[CourseEntry]:
        """
        Extract course information for a specific time slot (morning or afternoon).
        """
        courses = []

        if "Day" not in column_data_dict or time_key not in column_data_dict:
            return courses

        # Skip the very first row if it contains headers
        start_idx = 0
        if column_data_dict["Day"] and str(column_data_dict["Day"][0]).lower() == "day":
            start_idx = 1

        for i in range(start_idx, len(column_data_dict["Day"])):
            raw_course_code = column_data_dict[time_key][i]
            if raw_course_code is None:
                continue

            course_code = str(raw_course_code).strip()
            if not course_code or course_code.lower() == "none":
                continue

            if "8.30" in course_code.lower() or "1.30" in course_code.lower():
                continue

            if course_code.upper() in ["COURSE", "COURSES", "DAY", "CAMPUS"]:
                continue

            day_val = str(column_data_dict["Day"][i] or "")
            suffix = "_Afternoon" if "_Afternoon" in time_key else ""

            start_time = parse_exam_datetime(day_val, start_time_str, self.timezone)
            end_time = parse_exam_datetime(day_val, end_time_str, self.timezone)

            if not start_time or not end_time:
                self.logger.warning(f"Failed to parse datetime for {course_code} on {day_val}")
                continue

            course_info = CourseEntry(
                course_code=course_code,
                start_time=start_time,
                end_time=end_time,
                venue=str(column_data_dict.get(f"Venue{suffix}", [""] * (i + 1))[i] or "").strip(),
                coordinator=str(column_data_dict.get("Coordinator", [""] * (i + 1))[i] or "").strip(),
                hrs=str(column_data_dict.get(f"Hours{suffix}", [""] * (i + 1))[i] or "").strip(),
                raw_data={
                    "original_day": day_val,
                    "campus": column_data_dict.get("Campus", [""] * (i + 1))[i],
                    "invigilator": column_data_dict.get(f"Invigilators{suffix}", [""] * (i + 1))[i],
                }
            )

            courses.append(course_info)
        return courses
