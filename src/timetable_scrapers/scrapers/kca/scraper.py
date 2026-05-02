import re
from datetime import datetime
from typing import Any, List, Optional
from openpyxl import load_workbook

from ...base.scraper import BaseTimetableScraper
from ...registry import ScraperRegistry
from ...schemas import CourseEntry
from ...utils.time_parser import parse_exam_datetime


@ScraperRegistry.register("kca")
class KCAScraper(BaseTimetableScraper):
    """
    Scraper for KCA University exam timetable format.
    Matches the minimal Professor API contract.
    """

    @property
    def institution_name(self) -> str:
        return "KCA University"

    @property
    def timezone(self) -> str:
        return "EAT"  # East Africa Time

    def extract(self, file) -> List[CourseEntry]:
        """Extract exam timetable data from KCA Excel file."""
        wb_obj = load_workbook(file, data_only=True)

        key_map = {
            "DATE": "DATE",
            "TIME": "TIME",
            "VENUE": "ROOM|VENUE",
            "UNIT_CODE": "UNIT_CODE|UNIT CODE",
            "CAMPUS": "CAMPUS",
            "COORDINATOR": "INVIGILATOR OF THE DAY",
            "INVIGILATOR": "INVIGILATORS|PRINCIPAL INVIGILATORS|INVIGILATOR",
        }
        unit_code_pattern = re.compile(r"^[A-Z]{2,4}\s*\d{3,4}")

        all_courses: List[CourseEntry] = []

        for sheet_name in wb_obj.sheetnames:
            sheet = wb_obj[sheet_name]

            header_row: Optional[List[str]] = None
            header_idx = 0

            for row_idx, row in enumerate(
                sheet.iter_rows(values_only=True) if sheet else [], start=1
            ):
                if any(
                    "UNIT CODE" in str(cell).upper() for cell in row if cell
                ):
                    header_row = [
                        str(cell) if cell is not None else "" for cell in row
                    ]
                    header_idx = row_idx
                    break

            if not header_row:
                continue

            norm_headers = {
                self._normalize_header(h): idx
                for idx, h in enumerate(header_row)
                if h
            }

            for row_idx, row in enumerate(
                (
                    sheet.iter_rows(min_row=header_idx + 1, values_only=True)
                    if sheet
                    else []
                ),
                start=header_idx + 1,
            ):
                row = [cell if cell is not None else "" for cell in row]

                unit_code = ""
                for pattern in key_map["UNIT_CODE"].split("|"):
                    norm_pattern = self._normalize_header(pattern)
                    if norm_pattern in norm_headers:
                        unit_code = str(row[norm_headers[norm_pattern]]).strip()
                        break

                norm_uc = unit_code.upper().replace(" ", "").replace("_", "")
                if (
                    not unit_code
                    or norm_uc
                    in {
                        "UNITCODE",
                        "UNIT",
                        "TIME",
                        "DATE",
                        "DAY",
                        "VENUE",
                        "ROOM",
                        "EXAMINATION",
                    }
                    or not unit_code_pattern.search(unit_code.upper())
                ):
                    continue

                raw_data = {
                    "campus": "",
                    "invigilator": "",
                    "original_date": "",
                    "original_time": "",
                }

                date_val = ""
                time_range_str = ""
                venue_val = ""
                coordinator_val = ""

                for out_key, patterns in key_map.items():
                    for pattern in patterns.split("|"):
                        norm_pattern = self._normalize_header(pattern)
                        if norm_pattern in norm_headers:
                            val = row[norm_headers[norm_pattern]]
                            if out_key == "DATE":
                                date_val = self._convert_date(val)
                                raw_data["original_date"] = date_val
                            elif out_key == "TIME":
                                time_range_str = str(val).strip()
                                raw_data["original_time"] = time_range_str
                            elif out_key == "VENUE":
                                venue_val = str(val).strip()
                            elif out_key == "CAMPUS":
                                raw_data["campus"] = str(val).strip()
                            elif out_key == "COORDINATOR":
                                coordinator_val = str(val).strip()
                            elif out_key == "INVIGILATOR":
                                raw_data["invigilator"] = str(val).strip()
                            break

                # Split time range and parse
                start_iso = ""
                end_iso = ""
                if "-" in time_range_str:
                    time_parts = time_range_str.split("-", 1)
                    start_iso = parse_exam_datetime(date_val, time_parts[0].strip(), self.timezone)
                    end_iso = parse_exam_datetime(date_val, time_parts[1].strip(), self.timezone)

                if not start_iso or not end_iso:
                    continue

                all_courses.append(CourseEntry(
                    course_code=unit_code,
                    start_time=start_iso,
                    end_time=end_iso,
                    venue=venue_val or "TBA",
                    coordinator=coordinator_val,
                    raw_data=raw_data,
                ))

        return self.normalize_output(all_courses)

    def _convert_date(self, date_val: Any) -> str:
        """Convert Excel serial or string to readable date."""
        if isinstance(date_val, (int, float)):
            try:
                return datetime.fromordinal(
                    datetime(1900, 1, 1).toordinal() + int(date_val) - 2
                ).strftime("%Y-%m-%d")
            except ValueError:
                return str(date_val)
        return str(date_val).strip()

    def _normalize_header(self, header_str: str) -> str:
        """Normalize header string for matching."""
        if not header_str:
            return ""
        normalized = str(header_str).upper().strip()
        normalized = re.sub(r"[^\w\s]", "", normalized)
        normalized = re.sub(r"\s+", "_", normalized)
        return normalized
