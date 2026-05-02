import re
from typing import List
from openpyxl import load_workbook

from ...base.scraper import BaseTimetableScraper
from ...registry import ScraperRegistry
from ...schemas import CourseEntry
from ...utils.time_parser import parse_exam_datetime


@ScraperRegistry.register("strath")
class StrathScraper(BaseTimetableScraper):
    """
    Scraper for Strathmore University Exams.
    Matches the minimal Professor API contract.
    Handles merged cells by tracking current values across rows.
    """

    @property
    def institution_name(self) -> str:
        return "Strathmore University"

    @property
    def timezone(self) -> str:
        return "EAT"  # East Africa Time

    def extract(self, file) -> List[CourseEntry]:
        """
        Extract timetable data from Strathmore Excel file.
        """
        wb_obj = load_workbook(file, data_only=True, read_only=True)
        all_courses = []

        for sheet in wb_obj.worksheets:
            col_map = {
                "date": 0,
                "time": 2,
                "course": 5,
                "group": 6,
                "number": 7,
                "venue": 8,
                "lecturer": 9,
            }

            header_row_idx = -1

            # Try to find header row and detect columns
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True, max_row=10)):
                row_vals = [str(cell).strip().lower() if cell else "" for cell in row]
                if "time" in row_vals and any(word in "".join(row_vals) for word in ["course", "unit", "examination", "subject"]):
                    header_row_idx = row_idx
                    for i, val in enumerate(row_vals):
                        val_str = str(val).lower()
                        if "date" in val_str: col_map["date"] = i
                        elif "time" in val_str: col_map["time"] = i
                        elif any(kw in val_str for kw in ["unit", "course", "examination"]): col_map["course"] = i
                        elif "group" in val_str: col_map["group"] = i
                        elif "no" in val_str or "number" in val_str: col_map["number"] = i
                        elif "venue" in val_str: col_map["venue"] = i
                        elif any(kw in val_str for kw in ["invigilator", "lecturer"]): col_map["lecturer"] = i
                    break

            if header_row_idx == -1:
                continue

            current_date = ""
            current_time = ""
            current_course = ""
            current_group = ""
            current_number = ""
            current_venue = ""
            current_lecturer = ""

            start_row = header_row_idx + 1

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx < start_row:
                    continue

                if not any(row):
                    continue

                def get_val(idx):
                    return str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""

                date_val = get_val(col_map["date"])
                time_val = get_val(col_map["time"])
                course_val = get_val(col_map["course"])
                group_val = get_val(col_map["group"])
                number_val = get_val(col_map["number"])
                venue_val = get_val(col_map["venue"])
                lecturer_val = get_val(col_map["lecturer"])

                if date_val: current_date = date_val.rstrip(".")
                if time_val: current_time = time_val
                if course_val: current_course = course_val
                if group_val: current_group = group_val
                if number_val: current_number = number_val
                if venue_val: current_venue = venue_val
                if lecturer_val: current_lecturer = lecturer_val

                if current_course and current_group:
                    if course_val or group_val or venue_val or lecturer_val:
                        course_parts = current_course.split(":", 1)
                        course_code = course_parts[0].strip() if course_parts else current_course

                        # Parse time range (e.g., "8:00-10:00")
                        start_iso = ""
                        end_iso = ""
                        if "-" in current_time:
                            time_parts = current_time.split("-", 1)
                            start_iso = parse_exam_datetime(current_date, time_parts[0].strip(), self.timezone)
                            end_iso = parse_exam_datetime(current_date, time_parts[1].strip(), self.timezone)

                        if not start_iso or not end_iso:
                            continue

                        all_courses.append(CourseEntry(
                            course_code=course_code,
                            start_time=start_iso,
                            end_time=end_iso,
                            venue=current_venue or "TBA",
                            coordinator=current_lecturer,
                            raw_data={
                                "course_name": course_parts[1].strip() if len(course_parts) > 1 else "",
                                "group": current_group,
                                "student_count": current_number,
                                "original_date": current_date,
                                "original_time": current_time,
                            },
                        ))

        return self.normalize_output(all_courses)
