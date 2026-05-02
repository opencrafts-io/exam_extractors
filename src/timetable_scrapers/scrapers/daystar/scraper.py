from datetime import datetime
from typing import List
from openpyxl import load_workbook

from ...base.scraper import BaseTimetableScraper
from ...registry import ScraperRegistry
from ...schemas import CourseEntry
from ...utils.time_parser import parse_exam_datetime


@ScraperRegistry.register("school_exams")
class SchoolExamScraper(BaseTimetableScraper):
    """
    Scraper for Daystar University Exams.
    Matches the minimal Professor API contract.
    Handles multiple worksheets, each with room information in the first column.
    """

    @property
    def institution_name(self) -> str:
        return "Daystar University"

    @property
    def timezone(self) -> str:
        return "EAT"  # East Africa Time

    def extract(self, file) -> List[CourseEntry]:
        """Extract exam timetable data from school Excel file."""
        wb_obj = load_workbook(filename=file)
        work_sheets = wb_obj.sheetnames

        rooms = {}
        courses = []

        days_of_the_week = [
            "MONDAY",
            "TUESDAY",
            "WEDNESDAY",
            "THURSDAY",
            "FRIDAY",
            "SATURDAY",
        ]

        for sheet in work_sheets:
            work_sheet = wb_obj[sheet]

            for column_one in work_sheet.iter_cols(values_only=True):
                for i, room in enumerate(column_one):
                    if room is None or room == "ROOM":
                        continue
                    rooms[f"{i}"] = room
                break

            data_columns = list(work_sheet.iter_cols(values_only=True))[1:]

            day = ""
            course_time_range = ""
            course_code = ""

            for column in data_columns:
                for idx, value in enumerate(column):
                    if value is None:
                        continue

                    if value == "CHAPEL":
                        continue

                    if isinstance(value, datetime):
                        day = value.strftime("%Y-%m-%d")
                    elif (
                        isinstance(value, str)
                        and any(d in value.upper() for d in days_of_the_week)
                    ):
                        day = value

                    elif (
                        isinstance(value, str)
                        and len(value) > 0
                        and value[0].isdigit()
                    ):
                        course_time_range = value.strip()
                    elif isinstance(value, str):
                        course_code = value

                        # Parse time range and create entry
                        start_iso = ""
                        end_iso = ""
                        if "-" in course_time_range:
                            time_parts = course_time_range.split("-", 1)
                            start_iso = parse_exam_datetime(day, time_parts[0].strip(), self.timezone)
                            end_iso = parse_exam_datetime(day, time_parts[1].strip(), self.timezone)

                        if not start_iso or not end_iso:
                            continue

                        courses.append(
                            CourseEntry(
                                course_code=course_code,
                                start_time=start_iso,
                                end_time=end_iso,
                                venue=str(rooms.get(f"{idx}", "")).strip() or "TBA",
                                raw_data={
                                    "original_day": day,
                                    "original_time": course_time_range,
                                }
                            )
                        )

        return self.normalize_output(courses)
